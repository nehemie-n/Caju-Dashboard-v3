#!/usr/bin/env python3

import argparse
import os
import sys
import django
import openpyxl
from dotenv import load_dotenv

load_dotenv()

BASE_DIR = os.path.dirname(os.path.realpath(__name__))
sys.path.append(BASE_DIR)
os.environ.setdefault(
    "DJANGO_SETTINGS_MODULE", "cajulab_remote_sensing_dashboard.settings"
)
django.setup()

from apps.dashboard import models as dash_models
from django.db import connections


def add_countries(db):
    wb_obj = openpyxl.load_workbook(
        os.path.join(BASE_DIR, "scripts/setup_scripts/countries_with_coord.xlsx")
    )
    sheet = wb_obj.active
    for i in range(2, sheet.max_row + 1):
        country_name = country_name = sheet.cell(row=i, column=4).value
        print(f"Saving {country_name} in {db}")
        try:
            dash_models.Country.objects.using(db).get(country_name=country_name)
        except:
            country_elmt = dash_models.Country(
                country_code=sheet.cell(row=i, column=1).value,
                latitude=sheet.cell(row=i, column=2).value,
                longitude=sheet.cell(row=i, column=3).value,
                country_name=country_name,
                status=1,
            )
            country_elmt.save(using=db)


if __name__ == "__main__":
    for db in connections:
        add_countries(db=db)
